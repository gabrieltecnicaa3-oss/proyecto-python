import openpyxl, glob, os

# Buscar el archivo
pattern = r'C:\Users\usuar\OneDrive\Desktop\Info y Caracterizacion de Proveedores*.xlsx'
files = glob.glob(pattern)
print("Archivos encontrados:", files)
if not files:
    raise FileNotFoundError("No se encontró el archivo")

path = files[0]
wb = openpyxl.load_workbook(path, data_only=True)
print("Hojas:", wb.sheetnames)
ws = wb.active
print(f"Hoja activa: {ws.title}")

# Listar primeras 15 filas de la hoja activa
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 15:
        break
    print(row)
wb.close()
